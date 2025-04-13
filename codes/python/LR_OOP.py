import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
import statsmodels.api as sm
from scipy import stats
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from scipy.stats import shapiro
from statsmodels.stats.diagnostic import het_breuschpagan
from statsmodels.stats.stattools import durbin_watson

from codes.python.execution_timer import measure_execution_time, append_execution_time

class DataLoader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.data = None

    def load_data(self):
        try:
            self.data = pd.read_excel(self.file_path)
            print(f"Data loaded successfully from {self.file_path}")
        except Exception as e:
            print(f"Error loading data: {e}")
        return self.data

    def preprocess_data(self):
        if self.data is not None:
            self.data = self.data.dropna()
        return self.data


class LinearRegression:
    def __init__(self, data, x_column, y_column):
        self.data = data
        self.x_column = x_column
        self.y_column = y_column
        self.results = None

    def perform_regression(self):
        try:
            x_values = self.data[self.x_column].values
            y_values = self.data[self.y_column].values

            x_mean = np.mean(x_values)
            y_mean = np.mean(y_values)
            x_diff_mean = x_values - x_mean
            y_diff_mean = y_values - y_mean

            slope = np.sum(x_diff_mean * y_diff_mean) / np.sum(x_diff_mean ** 2)
            intercept = y_mean - slope * x_mean
            predictions = slope * x_values + intercept
            residuals = y_values - predictions

            ssm = np.sum((predictions - y_mean) ** 2)
            ssr = np.sum(residuals ** 2)
            sst = ssm + ssr

            num_samples = len(x_values)
            degrees_of_freedom = [1, num_samples - 2, num_samples - 1]

            msr = ssm / degrees_of_freedom[0]
            mse = ssr / degrees_of_freedom[1]

            f_statistic = msr / mse
            p_value = 1 - stats.f.cdf(f_statistic, degrees_of_freedom[0], degrees_of_freedom[1])

            r_squared = ssm / sst
            adjusted_r_squared = 1 - (1 - r_squared) * (num_samples - 1) / (num_samples - 2)
            r_value = np.sqrt(r_squared) if slope > 0 else -np.sqrt(r_squared)
            std_err = np.sqrt(mse / np.sum(x_diff_mean ** 2))
            rmse = np.sqrt(mse)
            mae = np.mean(np.abs(residuals))
            is_significant = p_value < 0.05

            self.results = {
                "slope": slope,
                "intercept": intercept,
                "r_value": r_value,
                "r_squared": r_squared,
                "adjusted_r_squared": adjusted_r_squared,
                "p_value": p_value,
                "std_err": std_err,
                "rmse": rmse,
                "mae": mae,
                "equation": f"y = {intercept:.4f} + {slope:.4f}x",
                "x_values": x_values,
                "y_values": y_values,
                "predictions": predictions,
                "residuals": residuals,
                "is_significant": is_significant,
                "ssm": ssm,
                "ssr": ssr,
                "sst": sst,
                "msr": msr,
                "mse": mse,
                "f_statistic": f_statistic,
                "degrees_of_freedom": degrees_of_freedom
            }
        except Exception as e:
            print(f"Error in regression calculations: {e}")
        return self.results

class RegressionTester:
    def __init__(self, regression):
        self.regression = regression
        self.residuals = regression.results["residuals"]
        self.x_values = regression.results["x_values"]

    def test_assumptions(self):
        # Homoscedasticity test (Breusch-Pagan)
        bp_test = het_breuschpagan(self.residuals, sm.add_constant(self.x_values))
        print(f"Breusch-Pagan test: p-value = {bp_test[1]:.4f}")

        # Normality test (Shapiro-Wilk)
        shapiro_test = shapiro(self.residuals)
        print(f"Shapiro-Wilk test: p-value = {shapiro_test.pvalue:.4f}")

        # Independence test (Durbin-Watson)
        dw_test = durbin_watson(self.residuals)
        print(f"Durbin-Watson test: statistic = {dw_test:.4f}")

class RegressionVisualizer:
    def __init__(self, regression, output_dir):
        self.regression = regression
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def create_plots(self):
        # Scatter plot
        plt.figure()
        sns.scatterplot(x=self.regression.results["x_values"], y=self.regression.results["y_values"])
        plt.plot(self.regression.results["x_values"], self.regression.results["predictions"], color="red")
        plt.title("Regression Scatter Plot")
        plt.xlabel("X")
        plt.ylabel("Y")
        plt.savefig(os.path.join(self.output_dir, "regression_scatter.png"))

        # Residual plot
        plt.figure()
        sns.residplot(x=self.regression.results["x_values"], y=self.regression.results["residuals"], lowess=True)
        plt.title("Residual Plot")
        plt.xlabel("X")
        plt.ylabel("Residuals")
        plt.savefig(os.path.join(self.output_dir, "residual_plot.png"))

        # Histogram of residuals
        plt.figure()
        sns.histplot(self.regression.results["residuals"], kde=True)
        plt.title("Histogram of Residuals")
        plt.xlabel("Residual Value")
        plt.ylabel("Frequency")
        plt.savefig(os.path.join(self.output_dir, "residual_histogram.png"))

        # QQ plot
        plt.figure()
        sm.qqplot(self.regression.results["residuals"], line="s")
        plt.title("QQ Plot")
        plt.savefig(os.path.join(self.output_dir, "qq_plot.png"))

class RegressionExporter:
    def __init__(self, regression, output_dir='../../output/LR/LR_OOP/excel', file_name='linear_regression_results.xlsx'):
        self.regression = regression
        self.results = regression.results
        self.output_dir = output_dir
        self.file_name = file_name
        os.makedirs(output_dir, exist_ok=True)

    def export_to_excel(self):
        if self.results is None:
            print("No results to export")
            return

        file_path = os.path.join(self.output_dir, self.file_name)

        # Create DataFrames
        df1 = self._create_computation_table()
        df2 = self._create_data_table()
        df3 = self._create_variability_table()

        # Write to Excel
        with pd.ExcelWriter(file_path) as writer:
            df1.to_excel(writer, sheet_name='Výpočtová tabuľka1', index=False)
            df2.to_excel(writer, sheet_name='Výpočtová tabuľka2', index=False)
            df3.to_excel(writer, sheet_name='LR', index=False)

        # Format the Excel file and add images
        self._format_excel_file(file_path)
        print(f"Data exported to {file_path}")

    def _create_computation_table(self):
        results = self.results
        yi = results['y_values']
        xi = results['x_values']
        xi_yi = xi * yi

        return pd.DataFrame({
            'yi': list(yi) + [np.sum(yi)],
            'xi': list(xi) + [np.sum(xi)],
            'xi*yi': list(xi_yi) + [np.sum(xi_yi)],
            'yi^2': list(yi ** 2) + [np.sum(yi ** 2)],
            'xi^2': list(xi ** 2) + [np.sum(xi ** 2)],
            'yi - y_mean': list(yi - np.mean(yi)) + [np.sum(yi - np.mean(yi))],
            'xi - x_mean': list(xi - np.mean(xi)) + [np.sum(xi - np.mean(xi))],
            '(yi - y_mean)*(xi - x_mean)': list((yi - np.mean(yi)) * (xi - np.mean(xi))) + [np.sum((yi - np.mean(yi)) * (xi - np.mean(xi)))],
            '(xi - x_mean)^2': list((xi - np.mean(xi)) ** 2) + [np.sum((xi - np.mean(xi)) ** 2)],
            '(yi - y_mean)^2': list((yi - np.mean(yi)) ** 2) + [np.sum((yi - np.mean(yi)) ** 2)],
        })

    def _create_data_table(self):
        results = self.results
        return pd.DataFrame({
            'Hodnoty závislej premennej yi': list(results['y_values']) + [np.sum(results['y_values'])],
            'Priemer (v m**2) xi': list(results['x_values']) + [np.sum(results['x_values'])],
            'y^i': list(results['predictions']) + [np.sum(results['predictions'])],
            'error': list(results['residuals']) + [np.sum(results['residuals'])],
            '(yi - y^i)^2': list(results['residuals'] ** 2) + [np.sum(results['residuals'] ** 2)],
        })

    def _create_variability_table(self):
        results = self.results
        variability_labels = ['Vysvetlená regresným modelom',
                              'Nevysvetlená regresným modelom',
                              'Celková variabilita']

        ss_values = ["SSM = {:.2f}".format(results['ssm']),
                     "SSR = {:.2f}".format(results['ssr']),
                     "SST = {:.2f}".format(results['sst'])]

        degrees_of_freedom = results['degrees_of_freedom']

        ms_values = ["MSA = {:.2f}".format(results['msr']),
                     "MSE = {:.2f}".format(results['mse']),
                     float('nan')]

        f_stat_values = ["F = {:.2f}".format(results['f_statistic']),
                         float('nan'),
                         float('nan')]

        p_values = ["P-value = {:.7f}".format(results['p_value']),
                    float('nan'),
                    float('nan')]

        significance_values = ["Áno" if results['is_significant'] else "Nie",
                               None,
                               None]

        return pd.DataFrame({
            'Variabilita premennej Y': variability_labels,
            'Súčet štvorcov SS': ss_values,
            'Stupne voľnosti DF': degrees_of_freedom,
            'Priemerný súčet štvorcov MS': ms_values,
            'F_stat': f_stat_values,
            'P-value': p_values,
            'Hladina významnosti': ["Hladina významnosti = {:.2f}".format(0.95),
                                     float('nan'),
                                     float('nan')],
            'ŠV modelu': significance_values
        })



    def _format_excel_file(self, file_path):
        workbook = load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if isinstance(cell.value, float):
                            cell.number_format = '0.0000'
                            current_length = len(str(int(cell.value))) + 5
                            if current_length > max_length:
                                max_length = current_length
                        elif cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

        image_paths = [
            os.path.join('../../output/LR/LR_OOP/graphs', 'regression_scatter.png'),
            os.path.join('../../output/LR/LR_OOP/graphs', 'residual_plot.png'),
            os.path.join('../../output/LR/LR_OOP/graphs', 'qq_plot.png'),
            os.path.join('../../output/LR/LR_OOP/graphs', 'residual_histogram.png')
        ]

        if any(os.path.exists(path) for path in image_paths):
            sheet = workbook.create_sheet("Graphs")
            row_pos = 1

            for i, img_path in enumerate(image_paths):
                if os.path.exists(img_path):
                    img = Image(img_path)
                    sheet.add_image(img, f'A{row_pos}')
                    row_pos += 20

        workbook.save(file_path)

def main():
    file_path = '../../input/LR/LR_100.xlsx'
    output_dir = "../../output/LR/LR_OOP/graphs"
    excel_output_dir = "../../output/LR/LR_OOP/excel"
    file_name = "linear_regression_results.xlsx"

    data_loader = DataLoader(file_path)
    data = data_loader.load_data()
    data = data_loader.preprocess_data()

    if data is None:
        print("No data to process")
        return

    regression = LinearRegression(data, "x", "y")
    results = regression.perform_regression()

    if results is None:
        print("Regression failed")
        return

    visualizer = RegressionVisualizer(regression, output_dir)
    visualizer.create_plots()

    tester = RegressionTester(regression)
    tester.test_assumptions()

    #exporter = RegressionExporter(regression, excel_output_dir, file_name)
    #exporter.export_to_excel()


if __name__ == "__main__":
    result = measure_execution_time(main)

    if isinstance(result, tuple):
        wait_time, execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")
        print(f"Waiting time: {wait_time:.6f} seconds")
        print(f"Active execution time: {execution_time - wait_time:.6f} seconds")

        append_execution_time(
            execution_time - wait_time,
            method="LR - OOP",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/h/moje/execution_times_python_small.xlsx"
        )
    else:
        execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")

        append_execution_time(
            execution_time,
            method="LR - OOP",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/h/moje/execution_times_python_small.xlsx"
        )