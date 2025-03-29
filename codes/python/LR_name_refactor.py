import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from scipy.stats import f
import matplotlib.pyplot as plt

from codes.python.execution_timer import measure_execution_time, timed_input


def load_data():
    data = pd.read_excel("../../input/mtcars/LR.xlsx")
    lr_data = data[['y', 'x']]
    lr_array = lr_data.to_numpy()
    sum_values = lr_array[-1, :]
    lr_array = lr_array[:-1, :]
    return lr_array, sum_values

def perform_calculations(array, sum_values):
    y_values = array[:, 0]
    x_values = array[:, 1]
    x_mean = np.mean(x_values)
    y_mean = np.mean(y_values)

    xy_product = x_values * y_values
    xy_mean = np.mean(xy_product)
    y_squared = y_values ** 2
    x_squared = x_values ** 2
    sum_y = sum_values[0]
    sum_x = sum_values[1]
    sum_x_squared = np.sum(x_values ** 2)
    sum_y_squared = np.sum(y_values ** 2)
    sum_xy = np.sum(x_values * y_values)
    x_diff_mean = x_values - x_mean
    y_diff_mean = y_values - y_mean
    y_diff_mean_squared = y_diff_mean ** 2
    x_diff_mean_squared = x_diff_mean ** 2
    y_diff_mean_squared_sum = np.sum(y_diff_mean_squared)
    x_diff_mean_squared_sum = np.sum(x_diff_mean_squared)
    num_samples = len(x_values)
    xy_diff_mean_product = (x_values - x_mean) * (y_values - y_mean)
    xy_diff_mean_product_sum = np.sum(xy_diff_mean_product)
    return x_mean, y_mean, xy_product, xy_mean, y_squared, x_squared, sum_y, sum_x, sum_x_squared, sum_y_squared, sum_xy, x_diff_mean, y_diff_mean, y_diff_mean_squared, x_diff_mean_squared, y_diff_mean_squared_sum, x_diff_mean_squared_sum, num_samples, xy_diff_mean_product, xy_diff_mean_product_sum

def calculate_regression_coefficients(covariance, scatter, y_mean, x_mean):
    b1 = covariance / scatter
    b0 = y_mean - (b1 * x_mean)
    return b0, b1

def calculate_covariance_and_scatter(xy_mean, x_mean, y_mean, x_squared):
    x_squared_mean = np.mean(x_squared)
    covariance = xy_mean - (x_mean * y_mean)
    scatter = x_squared_mean - (x_mean ** 2)
    return covariance, scatter

def calculate_balancing_line(b0, b1, x_values):
    balancing_line = b0 + (b1 * x_values)
    sum_balancing_line = np.sum(balancing_line)
    return balancing_line, sum_balancing_line

def calculate_squares_sum(y_values, y_mean, balancing_line):
    ssm = np.sum((balancing_line - y_mean) ** 2)
    ssr = np.sum((y_values - balancing_line) ** 2)
    sst = np.sum((y_values - y_mean) ** 2)
    return ssm, ssr, sst

def calculate_degrees_of_freedom(num_samples):
    return [1, num_samples - 2, num_samples - 1]

def calculate_mean_squares(ssr, ssm, degrees_of_freedom):
    msr = ssm / degrees_of_freedom[0]
    mse = ssr / degrees_of_freedom[1]
    return msr, mse

def validate_f_statistic(msm, mse, degrees_of_freedom):
    f_statistic = msm / mse
    user_input, wait_time = timed_input("Zadajte hodnotu hladiny významnosti (napr 0.95): ")
    significance_level = float(wait_time)
    p_value = f.sf(f_statistic, degrees_of_freedom[0], degrees_of_freedom[1])
    is_valid = f_statistic > f.ppf(significance_level, degrees_of_freedom[0], degrees_of_freedom[1]) and p_value < significance_level
    return is_valid, f_statistic, p_value, significance_level, wait_time


def write_data_to_excel(y_values, x_values, xy_product, y_squared, x_squared, x_diff_mean, y_diff_mean, xy_diff_mean_product, x_diff_mean_squared, y_diff_mean_squared, balancing_line, ssm, ssr, sst, degrees_of_freedom, msr, mse, f_statistic, p_value, significance_value, validation, file_name='output_data_LR.xlsx', output_dir='../../output/', image_path='../../output/LR_plot.png'):
    file_path = os.path.join(output_dir, file_name)
    error_values = y_values - balancing_line
    ssr_values = error_values ** 2
    df1 = pd.DataFrame({
        'Dom i': [i for i in range(len(y_values))] + ['∑'],
        'Spotreba (v kWh/mesiac) yi': list(y_values) + [np.sum(y_values)],
        'Priemer (v m**2) xi': list(x_values) + [np.sum(x_values)],
        'x*y': list(xy_product) + [np.sum(xy_product)],
        'y^2': list(y_squared) + [np.sum(y_squared)],
        'x^2': list(x_squared) + [np.sum(x_squared)],
        '(xi - x_avg)': list(x_diff_mean) + [np.sum(x_diff_mean_squared)],
        '(yi - y_avg)': list(y_diff_mean) + [np.sum(y_diff_mean)],
        '(xi - x_avg) - (yi - y_avg)': list(xy_diff_mean_product) + [np.sum(xy_diff_mean_product)],
        '(xi - x_avg)^2': list(x_diff_mean_squared) + [np.sum(x_diff_mean_squared)],
        '(yi - y_avg)^2': list(y_diff_mean_squared) + [np.sum(y_diff_mean_squared)],
    })

    df2 = pd.DataFrame({
        'Dom i': [i for i in range(len(y_values))] + ['∑'],
        'Spotreba (v kWh/mesiac) yi': list(y_values) + [np.sum(y_values)],
        'Priemer (v m**2) xi': list(x_values) + [np.sum(x_values)],
        'y^i': list(balancing_line) + [np.sum(balancing_line)],
        'error': list(error_values) + [np.sum(error_values)],
        '(yi - y^i)^2': list(ssr_values) + [np.sum(ssr_values)],
    })
    df3 = pd.DataFrame({
        'Variabilita premennej Y': ['Vysvetlená regresným modelom',
                                    'Nevysvetlená regresným modelom',
                                    'Celková variabilita'],

        'Súčet štvorcov SS': ["SSM = {:.2f}".format(ssm),
                              "SSR = {:.2f}".format(ssr),
                              "SST = {:.2f}".format(sst)],

        'Stupne voľnosti DF': degrees_of_freedom,

        'Priemerný súčet štvorcov MS': ["MSA = {:.2f}".format(msr),
                                        "MSE = {:.2f}".format(mse),
                                        float('nan')],
        'F_stat': [
            "F = {:.2f}".format(f_statistic),
            float('nan'),
            float('nan')],

        'P-value': [
            "P-value = {:.7f}".format(p_value),
            float('nan'),
            float('nan')],

        'Hladina významnosti': [
            "Hladina významnosti = {:.2f}".format(significance_value),
            float('nan'),
            float('nan')],

        'ŠV modelu': [
            "Áno" if validation and p_value < (1 -significance_value) else "Nie",
            None,
            None,
        ]
    })

    with pd.ExcelWriter(file_path) as writer:
        df1.to_excel(writer, sheet_name='Výpočtová tabuľka1', index=False)
        df2.to_excel(writer, sheet_name='Výpočtová tabuľka2', index=False)
        df3.to_excel(writer, sheet_name='LR', index=False)

    workbook = load_workbook(file_path)
    sheet1 = workbook['Výpočtová tabuľka1']
    sheet2 = workbook['Výpočtová tabuľka2']
    sheet3 = workbook['LR']

    for sheet in [sheet1, sheet2, sheet3]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        current_length = len(str(int(cell.value))) + 3
                        if current_length > max_length:
                            max_length = current_length
                    elif len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width
    if os.path.exists(image_path):
        sheet4 = workbook.create_sheet('Graf lineárnej regresie')
        img = Image(image_path)
        sheet4.add_image(img, 'A1')
    else:
        print(f"Warning: Image file not found at {image_path}, skipping image insertion")

    workbook.save(file_path)

    print(f"Data and image have been written to {file_path}")
    
def plot_regression_line(x_values, y_values, b0, b1, image_path='../../output/LR_plot.png'):

    if not os.path.exists('../../output'):
        os.makedirs('../../output')

    regression_line = b0 + b1 * x_values
    
    plt.figure(figsize=(10, 6))
    plt.scatter(x_values, y_values, color='blue', label='Reziduá')
    plt.plot(x_values, regression_line, color='red', label='Vyrovnávajúca priamka')

    for i in range(len(x_values)):
        plt.plot([x_values[i], x_values[i]], [y_values[i], regression_line[i]], color = "green", linestyle = 'dashed')

    plt.xlabel('X values')
    plt.ylabel('Y values')
    plt.title('Vyrovnávajúca priamka OLS')
    plt.legend()
    plt.savefig(image_path)
    
    plt.show()
    plt.close()

def main():
    wait_time = 0

    data_set, sum_values = load_data()
    y_values = data_set[:, 0]
    x_values = data_set[:, 1]
    x_mean, y_mean, xy_product, xy_mean, y_squared, x_squared, sum_y, sum_x, sum_x_squared, sum_y_squared, sum_xy, x_diff_mean, y_diff_mean, y_diff_mean_squared, x_diff_mean_squared, y_diff_mean_squared_sum, x_diff_mean_squared_sum, num_samples, xy_diff_mean_product, xy_diff_mean_product_sum = perform_calculations(data_set, sum_values)
    covariance, scatter = calculate_covariance_and_scatter(xy_mean, x_mean, y_mean, x_squared)
    b0, b1 = calculate_regression_coefficients(covariance, scatter, y_mean, x_mean)
    balancing_line, sum_balancing_line = calculate_balancing_line(b0, b1, data_set[:, 1])
    calculate_squares_sum(data_set[:, 0], y_mean, balancing_line)
    ssm, ssr, sst = calculate_squares_sum(data_set[:, 0], y_mean, balancing_line)
    degrees_of_freedom = calculate_degrees_of_freedom(num_samples)
    msr, mse = calculate_mean_squares(ssr, ssm, degrees_of_freedom)
    validation, f_statistic, p_value, significance_level, input_wait_time = validate_f_statistic(msr, mse, degrees_of_freedom)
    plot_regression_line(x_values, y_values, b0, b1)
    write_data_to_excel(y_values, x_values, xy_product, y_squared, x_squared, x_diff_mean, y_diff_mean, xy_diff_mean_product, x_diff_mean_squared, y_diff_mean_squared, balancing_line, ssm, ssr, sst, degrees_of_freedom, msr, mse, f_statistic, p_value, significance_level, validation)

    wait_time += input_wait_time
    return wait_time


if __name__ == '__main__':
    result = measure_execution_time(main)

    if isinstance(result, tuple):
        wait_time, execution_time = result
        print(f"Total execution time: {execution_time:.6f} seconds")
        print(f"Waiting time: {wait_time:.6f} seconds")
        print(f"Active execution time: {execution_time - wait_time:.6f} seconds")
    else:
        execution_time = result
        print(f"Total execution time: {execution_time:.6f} seconds")
