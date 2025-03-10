import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.drawing.image import Image
from scipy.stats import f



def without_nan(array):
    suma = []
    pocty = []
    bez_nan = []

    for row in array:
        valid_values = [i for i in row if not isinstance(i, float) or not i != i]
        bez_nan.extend(valid_values)
        sumy = sum(i for i in row if not isinstance(i, float) or not i != i)
        suma.append(sumy)
        pocty.append(len(valid_values))

    bez_nan = [float(i) for i in bez_nan]
    suma_bez_nan = np.sum(~np.isnan(array))

    return bez_nan, suma, pocty, suma_bez_nan


def priemer_and_priemer_c(count, suma):

    priemer = [suma[i] / pocty[i] for i in range(len(count))]
    #toto asi treba zmenit :D, resp. zmenit zapis do excelu, lebo to je matuce
    priemer_c = float(sum(suma) / sum(count))

    return priemer, priemer_c

def SSA(count, priemer, priemer_c):
    SSA_values = []

    for i in range(len(count)):
        tmp = ((priemer[i]- priemer_c) **2) * count[i]
        SSA_values.append(tmp)

    SSA_values = [float(i) for i in SSA_values]
    SSA_value = sum(SSA_values)

    return SSA_values, SSA_value

def SSE(count, priemer, bez_nan):
    start_index = 0
    SSE_values = []

    for i in range(len(count)):
        end_index = start_index + count[i]
        group_values = bez_nan[start_index:end_index]
        tmp = sum((value - priemer[i]) **2 for value in group_values)
        SSE_values.append(tmp)
        start_index = end_index

    SSE_values = [float(i) for i in SSE_values]
    SSE_value = sum(SSE_values)

    return SSE_values, SSE_value

def SST(count, bez_nan, priemer_c):
    start_index = 0
    SST_values = []

    for i in range(len(count)):

        end_index = start_index + count[i]
        group_values = bez_nan[start_index:end_index]
        tmp = sum((value - priemer_c) **2 for value in group_values)
        SST_values.append(tmp)
        start_index = end_index

        SST_values = [float(i) for i in SST_values]
    SST_value = sum(SST_values)

    return SST_values, SST_value

def df(pocty, sum_bez_nan):

    return [float(len(pocty) - 1), float(sum_bez_nan - (len(pocty))), float(sum_bez_nan - 1)]

def F_stat(SSA, SSE, count, sum_bez_nan):

    MSA = SSA / (len(count) - 1)
    MSE = SSE / (sum_bez_nan - (len(count)))
    F = MSA/MSE

    return MSA, MSE, F

def validate_f_statistic(degrees_of_freedom, f_statistic):

    significance_level = float(input("Zadajte hodnotu hladiny významnosti (napr. 0.95): "))
    p_value = f.sf(f_statistic, degrees_of_freedom[0], degrees_of_freedom[1])
    is_valid = f_statistic > f.ppf(significance_level, degrees_of_freedom[0], degrees_of_freedom[1]) and p_value < significance_level
    print(p_value, is_valid, significance_level)
    return is_valid, f_statistic, p_value, significance_level

#prepisat cely zapis do excelu, toto je bs
def write_data_to_excel_two_sheets(values_column, suma, pocty, priemer, priemer_c, ssa_value, ssa_values, sse_value, sse_values, sst_value, sst_values,
                                   df_values, MSA, MSE, F, p_value, significance_value, file_name='output_data.xlsx', output_dir='../../output/', image_path='../../output/ANOVA_plot.png'):

    file_path = os.path.join(output_dir, file_name)

    max_length = max(len(priemer) + 1, len(ssa_values) + 1, len(sse_values) + 1, len(values_column))

    priemer = priemer + [priemer_c] + [None] * (max_length - len(priemer) - 1)
    ssa_values = ssa_values + [ssa_value] + [None] * (max_length - len(ssa_values) - 1)
    sse_values = sse_values + [sse_value] + [None] * (max_length - len(sse_values) - 1)
    sst_values = sst_values + [sst_value] + [None] * (max_length - len(sst_values) - 1)
    values_column = values_column + [None] * (max_length - len(values_column))

    df1 = pd.DataFrame({
        'Priemer': priemer,
        'Hodnoty predajní': values_column,
        'SSA hodnoty': [value if value is not None else None for value in ssa_values],
        'SSE Values': [value if value is not None else None for value in sse_values],
        'SST Values': [value if value is not None else None for value in sst_values]
    })

    df2 = pd.DataFrame({
        'Zdroj variability premennej Y': ['Faktor', 'Nahoda', 'Spolu'],

        'Súčet štvorcov odchýlok SS': [
            "SSA = {:.2f}".format(ssa_value),
            "SSE = {:.2f}".format(sse_value),
            "SST = {:.2f}".format(sst_value)],

        'Počet stupňov voľnosti df': df_values,

        'Priemerný súčet štvorcov MS': [
            "MSA = {:.2f}".format(MSA),
            "MSE = {:.2f}".format(MSE),
            float('nan')],

        'F_stat': [
            "F = {:.2f}".format(F),
            float('nan'),
            float('nan')],

        'P-value': [
            "P-value = {:.7f}".format(p_value),
            float('nan'),
            float('nan')],

        'Hladina významnosti': [
            "Hladina významnosti = {:.2f}".format(significance_value),
            float('nan'),
            float('nan')],

        'ŠV modelu': [
            "Áno" if validation and p_value < (1 -significance_value) else "Nie",
            None,
            None,
        ]
    })

    with pd.ExcelWriter(file_path) as writer:
        df1.to_excel(writer, sheet_name='Výpočtová tabuľka', index=False)
        df2.to_excel(writer, sheet_name='ANOVA', index=False)

    workbook = load_workbook(file_path)
    sheet1 = workbook['Výpočtová tabuľka']
    sheet2 = workbook['ANOVA']

    for cell in sheet1['B']:
        cell.alignment = Alignment(horizontal='center')

    number_format = NamedStyle
    for sheet in [sheet1, sheet2]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        current_length = len(str(int(cell.value))) + 3
                        if current_length > max_length:
                            max_length = current_length
                    elif len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

    if os.path.exists(image_path):
        sheet3 = workbook.create_sheet('Boxplot')
        img = Image(image_path)
        sheet3.add_image(img, 'A1')
    else:
        print(f"Warning: Image file not found at {image_path}, skipping image insertion")

    workbook.save(file_path)

    print(f"Data and image have been written to {file_path}")


def boxplot(values_column, priemer, image_path='../../output/ANOVA_plot.png'):

    if not os.path.exists('../../output/'):
        os.makedirs('../../output/')
    data = [list(map(float, row.split(','))) for row in values_column]

    plt.figure(figsize=(10, 6))
    sns.boxplot(data=data)

    plt.title('Boxplot', fontsize=14)
    plt.xlabel('Store', fontsize=12)
    plt.ylabel('Value', fontsize=12)

    for i, p in enumerate(priemer):
        plt.axhline(p, color='r', linestyle='--', label=f'Priemer {i + 1} = {p:.2f}')

    plt.legend(title='Priemer', fontsize=12)
    plt.xticks(ticks=range(len(data)), labels=[f'Store {i + 1}' for i in range(len(data))])
    plt.savefig(image_path)
    plt.show()

data = pd.read_excel("../../input/mtcars/tst.xlsx")

P_data = data[['Predajňa 1', 'Predajňa 2', 'Predajňa 3']]
P_array = P_data.to_numpy()

bez_nan, suma, pocty, suma_bez_nan = without_nan(P_array)

priemer, priemer_c = priemer_and_priemer_c(pocty, suma)
sum_bez_nan = np.sum(~np.isnan(P_array))

SSA_values, SSA_value = SSA(pocty, priemer, priemer_c)
SSE_values, SSE_value = SSE(pocty, priemer, bez_nan)
SST_values, SST_value = SST(pocty, bez_nan, priemer_c)
df = df(pocty, sum_bez_nan)
MSA, MSE, F = F_stat(SSA_value, SSE_value, pocty, sum_bez_nan)
validation, f_statistic, p_value, significance_level = validate_f_statistic(df, F)

values_column = [','.join(map(str, row)) for row in P_array]

boxplot(values_column, priemer)
write_data_to_excel_two_sheets(values_column, suma, pocty, priemer, priemer_c, SSA_value, SSA_values, SSE_value, SSE_values, SST_value, SST_values, df, MSA, MSE, F, p_value, significance_level)









