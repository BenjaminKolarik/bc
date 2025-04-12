import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import statsmodels.api as sm
import os
from scipy import stats
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from scipy.stats import shapiro
from statsmodels.stats.diagnostic import het_breuschpagan
from statsmodels.stats.stattools import durbin_watson

from codes.python.execution_timer import measure_execution_time, append_execution_time, timed_input

def data_load(file_path):

    try:
        data = pd.read_excel(file_path)
        return data
    except Exception as e:
        print(f"Error loading data: {e}")
        return None

def preprocess_data(data):

    if data is None:
        return None
    else:
        data = data.dropna()
        return data


def perform_calculations(x_values, y_values):

    num_samples = len(x_values)

    # Calculate means
    x_mean = np.mean(x_values)
    y_mean = np.mean(y_values)

    # Calculate products and squares
    xy_product = x_values * y_values
    xy_mean = np.mean(xy_product)
    y_squared = y_values ** 2
    x_squared = x_values ** 2

    # Calculate sums
    sum_y = np.sum(y_values)
    sum_x = np.sum(x_values)
    sum_x_squared = np.sum(x_squared)
    sum_y_squared = np.sum(y_squared)
    sum_xy = np.sum(xy_product)

    # Calculate differences from mean
    x_diff_mean = x_values - x_mean
    y_diff_mean = y_values - y_mean

    # Calculate squared differences
    y_diff_mean_squared = y_diff_mean ** 2
    x_diff_mean_squared = x_diff_mean ** 2

    # Calculate sums of squared differences
    y_diff_mean_squared_sum = np.sum(y_diff_mean_squared)
    x_diff_mean_squared_sum = np.sum(x_diff_mean_squared)

    # Calculate product of differences
    xy_diff_mean_product = x_diff_mean * y_diff_mean
    xy_diff_mean_product_sum = np.sum(xy_diff_mean_product)

    return {
        'x_mean': x_mean,
        'y_mean': y_mean,
        'xy_product': xy_product,
        'xy_mean': xy_mean,
        'y_squared': y_squared,
        'x_squared': x_squared,
        'sum_y': sum_y,
        'sum_x': sum_x,
        'sum_x_squared': sum_x_squared,
        'sum_y_squared': sum_y_squared,
        'sum_xy': sum_xy,
        'x_diff_mean': x_diff_mean,
        'y_diff_mean': y_diff_mean,
        'y_diff_mean_squared': y_diff_mean_squared,
        'x_diff_mean_squared': x_diff_mean_squared,
        'y_diff_mean_squared_sum': y_diff_mean_squared_sum,
        'x_diff_mean_squared_sum': x_diff_mean_squared_sum,
        'num_samples': num_samples,
        'xy_diff_mean_product': xy_diff_mean_product,
        'xy_diff_mean_product_sum': xy_diff_mean_product_sum
    }


def calculate_covariance_and_scatter(xy_mean, x_mean, y_mean, x_squared_mean, x_mean_squared):

    covariance = xy_mean - x_mean * y_mean
    scatter = x_squared_mean - x_mean_squared
    return covariance, scatter


def calculate_regression_coefficients(covariance, scatter, y_mean, x_mean):

    b1 = covariance / scatter
    b0 = y_mean - b1 * x_mean
    return b0, b1


def calculate_balancing_line(b0, b1, x_values):

    balancing_line = b0 + b1 * x_values
    sum_balancing_line = np.sum(balancing_line)
    return balancing_line, sum_balancing_line


def calculate_squares_sum(y_values, y_mean, balancing_line):

    error_values = y_values - balancing_line

    ssr_values = error_values ** 2
    ssr = np.sum(ssr_values)

    ssm_values = (balancing_line - y_mean) ** 2
    ssm = np.sum(ssm_values)

    sst_values = (y_values - y_mean) ** 2
    sst = np.sum(sst_values)

    if not np.isclose(sst, ssr + ssm, rtol=1e-10):
        print(f"Warning: Sum of squares identity not satisfied. SST={sst}, SSR+SSM={ssr + ssm}")

    return ssm, ssr, sst, error_values, ssr_values


def calculate_degrees_of_freedom(num_samples):

    df_model = 1
    df_residual = num_samples - 2
    df_total = num_samples - 1

    return [df_model, df_residual, df_total]


def calculate_mean_squares(ssr, ssm, degrees_of_freedom):

    msr = ssm / degrees_of_freedom[0]
    mse = ssr / degrees_of_freedom[1]

    return msr, mse


def calculate_f_statistic(msr, mse, df_model, df_residual):

    f_statistic = msr / mse

    p_value = 1 - stats.f.cdf(f_statistic, df_model, df_residual)

    return f_statistic, p_value

def validate_significance(p_value, default_significance=0.95):
    significance_value = default_significance
    is_significant = p_value < significance_value

    print(f"Significance level: {significance_value}")
    print(f"P-value: {p_value}")
    print(f"Result is {'significant' if is_significant else 'not significant'} at the {significance_value} level.")
    #user_input, wait_time = timed_input("\nDo you want to change the significance level? (y/n): ")

    # if user_input.lower() == 'y':
    #     try:
    #         sig_input, sig_wait_time = timed_input("Enter the significance level (e.g., 0.90): ")
    #         wait_time += sig_wait_time
    #         new_significance = float(sig_input)
    #         is_significant = p_value < new_significance
    #         print(f"Result is {'significant' if is_significant else 'not significant'} at the {new_significance} level.")
    #     except ValueError:
    #         print("Invalid input. Using default significance level.")
    return is_significant, significance_value #wait_time



def perform_linear_regression(data, x_column, y_column):
    try:
        x_values = data[x_column].values
        y_values = data[y_column].values

        calculations = perform_calculations(x_values, y_values)

        x_squared_mean = np.mean(calculations['x_squared'])
        x_mean_squared = calculations['x_mean'] ** 2
        covariance, scatter = calculate_covariance_and_scatter(calculations['xy_mean'],calculations['x_mean'], calculations['y_mean'], x_squared_mean, x_mean_squared)

        b0, b1 = calculate_regression_coefficients(covariance, scatter, calculations['y_mean'], calculations['x_mean'])

        balancing_line, sum_balancing_line = calculate_balancing_line(b0, b1, x_values)

        ssm, ssr, sst, error_values, ssr_values = calculate_squares_sum(y_values, calculations['y_mean'], balancing_line)

        degrees_of_freedom = calculate_degrees_of_freedom(calculations['num_samples'])

        msr, mse = calculate_mean_squares(ssr, ssm, degrees_of_freedom)

        f_statistic, p_value = calculate_f_statistic(msr, mse, degrees_of_freedom[0], degrees_of_freedom[1])

        r_squared = ssm / sst
        adjusted_r_squared = 1 - (1 - r_squared) * (calculations['num_samples'] - 1) / (calculations['num_samples'] - 2)
        r_value = np.sqrt(r_squared) if b1 > 0 else -np.sqrt(r_squared)
        std_err = np.sqrt(mse / calculations['x_diff_mean_squared_sum'])
        rmse = np.sqrt(mse)
        mae = np.mean(np.abs(error_values))
        #sem by este malo ist wait_time ako premenna
        is_significant, significance_value = validate_significance(p_value)

        equation = f"y = {b0:.4f} + {b1:.4f}x"
        results = {
            "slope": b1,
            "intercept": b0,
            "r_value": r_value,
            "r_squared": r_squared,
            "adjusted_r_squared": adjusted_r_squared,
            "p_value": p_value,
            "std_err": std_err,
            "rmse": rmse,
            "mae": mae,
            "equation": equation,
            "x_values": x_values,
            "y_values": y_values,
            "predictions": balancing_line,
            "residuals": error_values,
            "is_significant": is_significant,
            "significance_value": significance_value,
            "ssm": ssm,
            "ssr": ssr,
            "sst": sst,
            "msr": msr,
            "mse": mse,
            "f_statistic": f_statistic,
            "degrees_of_freedom": degrees_of_freedom
        }

        return results#, wait_time

    except Exception as e:
        print(f"Error in linear regression calculations: {e}")
        import traceback
        traceback.print_exc()
        return None, 0

def test_assumptions(data, results):
    x_values = data['x']
    y_values = data['y']
    y_pred = results['predictions']
    residuals = results['residuals']

    # Homoscedasticity test (Breusch-Pagan)
    bp_test = het_breuschpagan(residuals, sm.add_constant(x_values))
    print(f"\nBreusch-Pagan test: p-value = {bp_test[1]:.4f}")

    # Normality test (Shapiro-Wilk)
    shapiro_test = shapiro(residuals)
    print(f"Shapiro-Wilk test: p-value = {shapiro_test.pvalue:.4f}")

    # Independence test (Durbin-Watson)
    dw_test = durbin_watson(residuals)
    print(f"Durbin-Watson test: statistic = {dw_test:.4f}")


def plot_regression_line(results, graph_dir):

    if not os.path.exists(graph_dir):
        os.makedirs(graph_dir)

    if results is None:
        print("No results found")
        return

    x_values = results['x_values']
    y_values = results['y_values']
    predictions = results['predictions']
    residuals = results['residuals']
    b0 = results['intercept']
    b1 = results['slope']

    image_path = os.path.join(graph_dir, "regression.png")

    plt.figure(figsize=(10, 6))
    plt.scatter(x_values, y_values, color = 'blue', label='Data Points')
    plt.plot(x_values, predictions, color = 'red', label='Linear Regression')

    plt.xlabel('x')
    plt.ylabel('y')
    plt.title(f'Linear Regression: y = {b0:.4f} + {b1:.4f}x')
    plt.legend()
    plt.savefig(image_path)
    plt.close()

    create_additional_graphs(results, graph_dir)

    return image_path

def create_additional_graphs(results, graph_dir):
    predictions = results['predictions']
    residuals = results['residuals']

    plt.figure(figsize=(10, 6))
    plt.scatter(predictions, residuals)
    plt.axhline(y=0, color='r', linestyle='-')
    plt.title('Residual Plot')
    plt.xlabel('Predicted Values')
    plt.ylabel('Residuals')
    plt.tight_layout()
    plt.savefig(os.path.join(graph_dir, 'residual_plot.png'))
    plt.close()

    plt.figure(figsize=(10, 6))
    stats.probplot(residuals, plot=plt)
    plt.title('Q-Q Plot of Residuals')
    plt.tight_layout()
    plt.savefig(os.path.join(graph_dir, 'qq_plot.png'))
    plt.close()

    plt.figure(figsize=(10, 6))
    sns.histplot(residuals, kde=True)
    plt.title('Histogram of Residuals')
    plt.xlabel('Residual Value')
    plt.ylabel('Frequency')
    plt.tight_layout()
    plt.savefig(os.path.join(graph_dir, 'residual_histogram.png'))
    plt.close()


def export_results_to_excel(results, excel_dir, file_name='linear_regression_results.xlsx'):

    if results is None:
        print("No results to export")
        return
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)

    file_path = os.path.join(excel_dir, file_name)

    x_values = results['x_values']
    y_values = results['y_values']
    xy_product = x_values * y_values
    y_squared = y_values ** 2
    x_squared = x_values ** 2
    x_mean = np.mean(x_values)
    y_mean = np.mean(y_values)

    x_diff_mean = x_values - x_mean
    y_diff_mean = y_values - y_mean
    xy_diff_mean_product = x_diff_mean * y_diff_mean
    x_diff_mean_squared = x_diff_mean ** 2
    y_diff_mean_squared = y_diff_mean ** 2

    balancing_line = results['predictions']
    error_values = results['residuals']
    ssr_values = error_values ** 2

    df1 = pd.DataFrame({
        'yi': list(y_values) + [np.sum(y_values)],
        'xi': list(x_values) + [np.sum(x_values)],
        'xi*yi': list(xy_product) + [np.sum(xy_product)],
        'yi^2': list(y_squared) + [np.sum(y_squared)],
        'xi^2': list(x_squared) + [np.sum(x_squared)],
        'yi - y_mean': list(y_diff_mean) + [np.sum(y_diff_mean)],
        'xi - x_mean': list(x_diff_mean) + [np.sum(x_diff_mean)],
        '(yi - y_mean)*(xi - x_mean)': list(xy_diff_mean_product) + [np.sum(xy_diff_mean_product)],
        '(xi - x_mean)^2': list(x_diff_mean_squared) + [np.sum(x_diff_mean_squared)],
        '(yi - y_mean)^2': list(y_diff_mean_squared) + [np.sum(y_diff_mean_squared)]
    })

    df2 = pd.DataFrame({
        'Hodnoty závislej premennej yi': list(y_values) + [np.sum(y_values)],
        'Priemer (v m**2) xi': list(x_values) + [np.sum(x_values)],
        'y^i': list(balancing_line) + [np.sum(balancing_line)],
        'error': list(error_values) + [np.sum(error_values)],
        '(yi - y^i)^2': list(ssr_values) + [np.sum(ssr_values)],
    })

    df3 = pd.DataFrame({
        'Variabilita premennej Y': ['Vysvetlená regresným modelom',
                                    'Nevysvetlená regresným modelom',
                                    'Celková variabilita'],

        'Súčet štvorcov SS': ["SSM = {:.2f}".format(results['ssm']),
                              "SSR = {:.2f}".format(results['ssr']),
                              "SST = {:.2f}".format(results['sst'])],

        'Stupne voľnosti DF': results['degrees_of_freedom'],

        'Priemerný súčet štvorcov MS': ["MSA = {:.2f}".format(results['msr']),
                                        "MSE = {:.2f}".format(results['mse']),
                                        float('nan')],

        'F_stat': [
            "F = {:.2f}".format(results['f_statistic']),
            float('nan'),
            float('nan')],

        'P-value': [
            "P-value = {:.7f}".format(results['p_value']),
            float('nan'),
            float('nan')],

        'Hladina významnosti': [
            "Hladina významnosti = {:.2f}".format(0.95),
            float('nan'),
            float('nan')],

        'ŠV modelu': [
            "Áno" if results['is_significant'] else "Nie",
            None,
            None,
        ]
    })

    with pd.ExcelWriter(file_path) as writer:
        df1.to_excel(writer, sheet_name='Výpočtová tabuľka1', index=False)
        df2.to_excel(writer, sheet_name='Výpočtová tabuľka2', index=False)
        df3.to_excel(writer, sheet_name='LR', index=False)

    format_excel_file(file_path, excel_dir.replace('excel', 'graphs'))
    print(f"\nData exported to {file_path}")


def format_excel_file(file_path, graph_dir):
    workbook = load_workbook(file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        current_length = len(str(int(cell.value))) + 3
                        if current_length > max_length:
                            max_length = current_length
                    elif cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

    image_paths = [
        os.path.join(graph_dir, 'regression.png'),
        os.path.join(graph_dir, 'residual_plot.png'),
        os.path.join(graph_dir, 'qq_plot.png'),
        os.path.join(graph_dir, 'residual_histogram.png')
    ]

    sheet4 = workbook.create_sheet('Graphs')
    row_pos = 1

    for i, img_path in enumerate(image_paths):
        if os.path.exists(img_path):
            img = Image(img_path)
            sheet4.add_image(img, f'A{row_pos}')
            row_pos += 40

    workbook.save(file_path)

def print_summary(results):
    if results is None:
        print("No results to summarize")
        return

    print("\nLinear Regression Summary:")
    print(f"Equation: {results['equation']}")
    print(f"R-squared: {results['r_squared']:.4f}")
    print(f"P-value: {results['p_value']:.4f}")
    print(f"Significant: {'Yes' if results['is_significant'] else 'No'}")
    print(f"RMSE: {results['rmse']:.4f}")
    print(f"MAE: {results['mae']:.4f}")
    print(f"F-statistic: {results['f_statistic']:.4f}")

def main():
    wait_time = 0
    data = data_load('../../input/LR/LR_100.xlsx')
    data = preprocess_data(data)

    if data is None:
        print("No data to process")
        return wait_time

    x_column = "x"
    y_column = "y"

    try: #tu by malo ist lr_wait_time ako druha premenna
        results = perform_linear_regression(data, x_column, y_column)
        #wait_time = lr_wait_time

        if results is None:
            print("Linear regression failed")
            return wait_time

        test_assumptions(data, results)
        print_summary(results)

        plot_regression_line(results, '../../output/LR/LR_base/graphs')
        export_results_to_excel(results, '../../output/LR/LR_base/excel')


    except Exception as e:
        print(f"Error in main function: {e}")
        import traceback
        traceback.print_exc()

    return wait_time


if __name__ == "__main__":
    result = measure_execution_time(main)

    if isinstance(result, tuple):
        wait_time, execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")
        print(f"Waiting time: {wait_time:.6f} seconds")
        print(f"Active execution time: {execution_time - wait_time:.6f} seconds")

        append_execution_time(
            execution_time - wait_time,
            method="LR - procedural - excel",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/execution_times_python_small.xlsx"
        )
    else:
        execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")

        append_execution_time(
            execution_time,
            method="LR - procedural - excel",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/execution_times_python_small.xlsx"
        )