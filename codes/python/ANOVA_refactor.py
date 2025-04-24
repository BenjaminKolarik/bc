import os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import scipy.stats as stats
from statsmodels.formula.api import ols
from statsmodels.stats.multicomp import pairwise_tukeyhsd

from codes.python.execution_timer import measure_execution_time, timed_input, append_execution_time

def data_load(file_path):

    return pd.read_excel(file_path)

def preprocess_data(data):
    data = data.dropna()
    data = data[data['value'] != 0]

    return data

def calculate_group_statistics(data_frame, group_col, value_col):
    groups = {}
    group_means = {}
    grand_mean = data_frame[value_col].mean()
    n_total = len(data_frame)

    for group_name, group_data in data_frame.groupby(group_col):
        groups[group_name] = group_data[value_col].values
        group_means[group_name] = group_data[value_col].mean()

    return groups, group_means, grand_mean, n_total

def calculate_sum_of_squares(groups, group_means, grand_mean):
    SSA = sum(len(groups[group_name]) * (group_means[group_name] - grand_mean) ** 2 for group_name in groups)
    SSE = sum(sum((value - group_means[group_name]) ** 2 for value in groups[group_name]) for group_name in groups)
    SST = SSA + SSE

    return SSA, SSE, SST


def calculate_degrees_of_freedom(groups):

    k = len(groups)
    n = sum(len(group_values) for group_values in groups.values())

    df_between = k - 1
    df_within = n - k
    df_total = n - 1

    return df_between, df_within, df_total

def calculate_mean_squares(SSA, SSE, df_between, df_within):
    MSA = SSA / df_between
    MSE= SSE / df_within

    return MSA, MSE

def calculate_f_statistic(MSA, MSE):

    return MSA / MSE

def calculate_p_value(f_statistic, df_between, df_within):

    return 1 - stats.f.cdf(f_statistic, df_between, df_within)

def validate_f_statistic(f_statistic, p_value, significance_value = 0.05):
    if significance_value is None:
        significance_value, wait_time = timed_input("Zadajte hladinu významnosti (napr. 0.95): ")
        significance_value = float(significance_value)
    else:
        wait_time = 0
    is_significant = p_value < significance_value
    return is_significant, significance_value, wait_time

def perform_anova(data_frame, group_col, value_col, significance_value = 0.05):
    wait_time = 0
    groups, group_means, grand_mean, n_total = calculate_group_statistics(data_frame, group_col, value_col)
    SSA, SSE, SST = calculate_sum_of_squares(groups, group_means, grand_mean)
    df_between, df_within, df_total = calculate_degrees_of_freedom(groups)
    MSA, MSE = calculate_mean_squares(SSA, SSE, df_between, df_within)
    f_statistic = calculate_f_statistic(MSA, MSE)
    p_value = calculate_p_value(f_statistic, df_between, df_within)
    is_significant, significance_value, input_wait_time = validate_f_statistic(f_statistic, p_value, significance_value)

    wait_time += input_wait_time

    results = {
        "groups": groups,
        "group_means": group_means,
        "grand_mean": grand_mean,
        "SSA": SSA,
        "SSE": SSE,
        "SST": SST,
        "df_between": df_between,
        "df_within": df_within,
        "df_total": df_total,
        "MSA": MSA,
        "MSE": MSE,
        "f_statistic": f_statistic,
        "p_value": p_value,
        "significance_value": significance_value,
        "is_significant": is_significant
    }

    return results, wait_time

def plot_anova_results(data_frame, group_col, value_col, output_dir = '../../output/ANOVA/ANOVA_base/graphs'):

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    #boxplot
    plt.figure(figsize=(10, 6))
    sns.boxplot(x = group_col, y = value_col, data = data_frame)
    plt.title('Boxplot of ANOVA Groups')
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'anova_boxplot.png'))

    #barplot
    plt.figure(figsize=(10, 6))
    sns.barplot(x = group_col, y = value_col, data = data_frame)
    plt.title('Barplot of ANOVA Groups')
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'anova_barplot.png'))

    plt.close('all')

def export_anova_results_to_excel(results, output_dir = '../../output/ANOVA/ANOVA_base/excel', file_name = 'anova_results.xlsx'):

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    file_path = os.path.join(output_dir, file_name)

    summary_df = pd.DataFrame({
        "Zdroj variability premennej Y": ["Faktor", "Náhoda", "Spolu"],
        "Suma štvorcov odchýlok SS": [results['SSA'], results['SSE'], results['SST']],
        "Počet stupňov voľnosti df": [results['df_between'], results['df_within'], results['df_total']],
        "Priemerný súčet štvorcov MS": [results['MSA'], results['MSE'], None],
        "F_stat": [results['f_statistic'], None, None],
        "P-value": [results['p_value'], None, None],
        "Hladina významnosti": [results['significance_value'], None, None],
        "ŠV modelu": ["Áno" if results['is_significant'] else "Nie", None, None]
    })

    group_means_df = pd.DataFrame({
        "Group": list(results['group_means'].keys()),
        "Mean": list(results['group_means'].values()),
        "Count": [len(results['groups'][group]) for group in results['groups']],
    })

    with pd.ExcelWriter(file_path) as writer:
        summary_df.to_excel(writer, sheet_name='ANOVA Summary', index=False)
        group_means_df.to_excel(writer, sheet_name='Group Statistics', index=False)

    workbook = load_workbook(file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if isinstance(cell.value, float):
                        cell.number_format = '0.00'
                        current_length = len(str(int(cell.value))) + 3
                        if current_length > max_length:
                            max_length = current_length
                    elif len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

    image_paths = [
        os.path.join('../../output/ANOVA/ANOVA_base/graphs', 'anova_boxplot.png'),
        #os.path.join('../../output/ANOVA/ANOVA_base/graphs', 'anova_residual_plot.png'),
        os.path.join('../../output/ANOVA/ANOVA_base/graphs', 'anova_barplot.png')
    ]

    if any(os.path.exists(path) for path in image_paths):
        sheet = workbook.create_sheet("Graphs")
        row_pos = 1

        for i, img_path in enumerate(image_paths):
            if os.path.exists(img_path):
                img = Image(img_path)
                sheet.add_image(img, f'A{row_pos}')
                row_pos += 20

    workbook.save(file_path)
    print(f"\nData and image have been written to {file_path}")

def test_assumptions(data_frame, group_col, value_col):
    model = ols(f'{value_col} ~ C({group_col})', data=data_frame).fit()
    residuals = model.resid

    shapiro_test = stats.shapiro(residuals)
    print(f"\nShapiro-Wilk test for normality: p-value = {shapiro_test.pvalue:.4f}")

    # Homogeneity of variances test (Levene's test)
    groups = [data_frame[data_frame[group_col] == group][value_col] for group in data_frame[group_col].unique()]
    levene_test = stats.levene(*groups)
    print(f"Levene's test for homogeneity of variances: p-value = {levene_test.pvalue:.4f}")

    tukey = pairwise_tukeyhsd(data_frame['value'], data_frame['group'], alpha = 0.05)
    print("\nPost-hoc analysis (Tukey's HSD):")
    print(tukey)


def main():
    wait_time = 0
    data = data_load('../../input/ANOVA/ANOVA_small.xlsx')
    data = preprocess_data(data)

    group_col = "group"
    value_col = "value"

    try:
        results, anova_wait_time = perform_anova(data, group_col, value_col)
        wait_time += anova_wait_time

        print(f"\nF-statistic: {results['f_statistic']:.4f}")
        print(f"P-value: {results['p_value']:.4f}")
        print(f"Significant: {'Yes' if results['is_significant'] else 'No'}")

        test_assumptions(data, group_col, value_col)

        plot_anova_results(data, group_col, value_col)
        export_anova_results_to_excel(results)

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

    return wait_time

if __name__ == "__main__":
    #name = timed_input("Name: ")
    result = measure_execution_time(main)

    if isinstance(result, tuple):
        wait_time, execution_time = result
        print(f"Total execution time: {execution_time:.6f} seconds")
        print(f"Waiting time: {wait_time:.6f} seconds")
        print(f"Active execution time: {execution_time - wait_time:.6f} seconds")

        append_execution_time(
            execution_time,
            method="ANOVA - procedural - excel",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/h/moje/execution_times_python_small.xlsx"
        )
    else:
        execution_time = result
        print(f"Total execution time: {execution_time:.6f} seconds")

        append_execution_time(
            execution_time,
            method="ANOVA - procedural - excel",
            computer_name="Windows Ryzen 9 5900x 32GB",
            excel_file="../../output/execution_times/h/moje/execution_times_python_small.xlsx"
        )


