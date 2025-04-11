import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import scipy.stats as stats
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from statsmodels.formula.api import ols

from codes.python.execution_timer import measure_execution_time, append_execution_time


class DataLoader:
    @staticmethod
    def load_data(file_path):
        try:
            data = pd.read_excel(file_path)
            print(f"Data loaded successfully from {file_path}")
            return data
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            return None

    @staticmethod
    def preprocess_data(data):
        if data is None:
            return None
        else:
            data = data.dropna()
            data = data[data['value'] != 0]

            return data


class ANOVA:
    def __init__(self, data, group_column, value_column, significance_level=0.95):
        self.data = data
        self.group_column = group_column
        self.value_column = value_column
        self.significance_level = significance_level
        self.results = None
        self.wait_time = 0

    def run_analysis(self):
        try:
            groups, group_means, grand_mean, n_total = self._calculate_group_statistics()
            SSA, SSE, SST = self._calculate_sum_of_squares(groups, group_means, grand_mean)
            df_between, df_within, df_total = self._calculate_degrees_of_freedom(groups)
            MSA, MSE = self._calculate_mean_squares(SSA, SSE, df_between, df_within)
            f_statistic = self._calculate_f_statistic(MSA, MSE)
            p_value = self._calculate_p_value(f_statistic, df_between, df_within)
            is_significant = p_value < self.significance_level

            self.results = {
                "groups": groups,
                "group_means": group_means,
                "grand_mean": grand_mean,
                "SSA": SSA,
                "SSE": SSE,
                "SST": SST,
                "df_between": df_between,
                "df_within": df_within,
                "df_total": df_total,
                "MSA": MSA,
                "MSE": MSE,
                "f_statistic": f_statistic,
                "p_value": p_value,
                "significance_value": self.significance_level,
                "is_significant": is_significant
            }

            return self.results

        except Exception as e:
            print(f"Error in ANOVA analysis: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _calculate_group_statistics(self):
        groups = {}
        group_means = {}

        for group_name, group_data in self.data.groupby(self.group_column):
            groups[group_name] = group_data[self.value_column].values
            group_means[group_name] = group_data[self.value_column].mean()

        grand_mean = self.data[self.value_column].mean()
        n_total = len(self.data)

        return groups, group_means, grand_mean, n_total

    def _calculate_sum_of_squares(self, groups, group_means, grand_mean):
        SSA = sum(len(group_values) * (group_means[group_name] - grand_mean)**2
                 for group_name, group_values in groups.items())

        SSE = sum(sum((value - group_means[group_name]) ** 2 for value in groups[group_name]) for group_name in groups)
        SST = SSA + SSE

        return SSA, SSE, SST

    def _calculate_degrees_of_freedom(self, groups):

        k = len(groups)

        n= sum(len(group_values) for group_values in groups.values())

        df_between = k - 1

        df_within = n - k

        df_total = n - 1

        return df_between, df_within, df_total

    def _calculate_mean_squares(self, SSA, SSE, df_between, df_within):

        MSA = SSA / df_between
        MSE = SSE / df_within

        return MSA, MSE

    def _calculate_f_statistic(self, MSA, MSE):

        return MSA/MSE

    def _calculate_p_value(self, f_statistic, df_between, df_within):
        p_value = 1 - stats.f.cdf(f_statistic, df_between, df_within)
        return p_value

    def get_results_summary(self):
        if self.results is None:
            return "Analysis not yet performed"

        return {
            "F-statistic": f"{self.results['f_statistic']:.4f}",
            "P-value": f"{self.results['p_value']:.4f}",
            "Significant": "Yes" if self.results['is_significant'] else "No",
            "Group means": self.results['group_means']
        }


class ANOVAVisualizer:
    def __init__(self, data, group_column, value_column, output_dir='../../output/ANOVA/ANOVA_OOP/graphs'):
        self.data = data
        self.group_column = group_column
        self.value_column = value_column
        self.output_dir = output_dir

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def create_plots(self):
        self._create_boxplot()
        self._create_barplot()
        plt.close('all')

    def _create_boxplot(self):
        plt.figure(figsize=(10, 6))
        sns.boxplot(x=self.group_column, y=self.value_column, data=self.data)
        plt.title('Boxplot of ANOVA Groups')
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'anova_boxplot.png'))

    def _create_barplot(self):
        plt.figure(figsize=(10, 6))
        sns.barplot(x=self.group_column, y=self.value_column, data=self.data)
        plt.title('Barplot of ANOVA Groups')
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'anova_barplot.png'))


class ANOVATester:
    def __init__(self, data, group_column, value_column):
        self.data = data
        self.group_column = group_column
        self.value_column = value_column

    def test_assumptions(self):
        model = ols(f'{self.value_column} ~ C({self.group_column})', data=self.data).fit()
        residuals = model.resid

        shapiro_test = stats.shapiro(residuals)
        print(f"\nShapiro-Wilk test for normality: p-value = {shapiro_test.pvalue:.4f}")

        # Homogeneity of variances test (Levene's test)
        groups = [self.data[self.data[self.group_column] == group][self.value_column] for group in self.data[self.group_column].unique()]
        levene_test = stats.levene(*groups)
        print(f"Levene's test for homogeneity of variances: p-value = {levene_test.pvalue:.4f}")


class ANOVAExporter:
    def __init__(self, results, output_dir='../../output/ANOVA/ANOVA_OOP/excel', file_name='anova_results.xlsx'):
        self.results = results
        self.output_dir = output_dir
        self.file_name = file_name

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def export_to_excel(self):
        if self.results is None:
            print("No results to export")
            return

        file_path = os.path.join(self.output_dir, self.file_name)

        summary_df = pd.DataFrame({
            "Zdroj variability premennej Y": ["Faktor", "Náhoda", "Spolu"],
            "Suma štvorcov odchýlok SS": [self.results['SSA'], self.results['SSE'], self.results['SST']],
            "Počet stupňov voľnosti df": [self.results['df_between'], self.results['df_within'],
                                          self.results['df_total']],
            "Priemerný súčet štvorcov MS": [self.results['MSA'], self.results['MSE'], None],
            "F_stat": [self.results['f_statistic'], None, None],
            "P-value": [self.results['p_value'], None, None],
            "Hladina významnosti": [self.results['significance_value'], None, None],
            "ŠV modelu": ["Áno" if self.results['is_significant'] else "Nie", None, None]
        })

        group_means_df = pd.DataFrame({
            "Group": list(self.results['group_means'].keys()),
            "Mean": list(self.results['group_means'].values()),
            "Count": [len(self.results['groups'][group]) for group in self.results['groups']],
        })

        with pd.ExcelWriter(file_path) as writer:
            summary_df.to_excel(writer, sheet_name='ANOVA Summary', index=False)
            group_means_df.to_excel(writer, sheet_name='Group Statistics', index=False)

        self._format_excel_file(file_path)
        print(f"Data exported to {file_path}")

    def _format_excel_file(self, file_path):
        workbook = load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if isinstance(cell.value, float):
                            cell.number_format = '0.00'
                            current_length = len(str(int(cell.value))) + 3
                            if current_length > max_length:
                                max_length = current_length
                        elif cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

        image_paths = [
            os.path.join('../../output/ANOVA/ANOVA_OOP/graphs', 'anova_boxplot.png'),
            os.path.join('../../output/ANOVA/ANOVA_OOP/graphs', 'anova_barplot.png')
        ]

        if any(os.path.exists(path) for path in image_paths):
            sheet = workbook.create_sheet("Graphs")
            row_pos = 1

            for i, img_path in enumerate(image_paths):
                if os.path.exists(img_path):
                    img = Image(img_path)
                    sheet.add_image(img, f'A{row_pos}')
                    row_pos += 20

        workbook.save(file_path)


def main():

    data_loader = DataLoader()
    file_path = '../../input/ANOVA/ANOVA_medium.xlsx'
    data = data_loader.load_data(file_path)
    data = data_loader.preprocess_data(data)

    if data is None:
        print("Failed to load or preprocess data")
        return

    group_col = "group"
    value_col = "value"

    if group_col not in data.columns or value_col not in data.columns:
        print(f"Columns '{group_col}' and/or '{value_col}' not found in the data")
        print(f"Available columns: {', '.join(data.columns)}")
        return

    try:
        anova = ANOVA(data, group_col, value_col)
        results = anova.run_analysis()

        if results is None:
            print("ANOVA analysis failed")
            return

        visualizer = ANOVAVisualizer(data, group_col, value_col)
        visualizer.create_plots()

        exporter = ANOVAExporter(results)
        exporter.export_to_excel()

        tester = ANOVATester(data, group_col, value_col)
        tester.test_assumptions()

        summary = anova.get_results_summary()
        print("\nANOVA Results:")
        print(f"F-statistic: {summary['F-statistic']}")
        print(f"P-value: {summary['P-value']}")
        print(f"Significant: {summary['Significant']}")

    except Exception as e:
        print(f"An error occurred in the ANOVA application: {str(e)}")
        import traceback
        traceback.print_exc()

    return


if __name__ == "__main__":
    result = measure_execution_time(main)

    if isinstance(result, tuple):
        wait_time, execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")
        print(f"Waiting time: {wait_time:.6f} seconds")
        print(f"Active execution time: {execution_time - wait_time:.6f} seconds")

        append_execution_time(
            execution_time - wait_time,
            method="ANOVA - OOP",
            computer_name="Windows Ryzen 9 5900x 32GB"
        )
    else:
        execution_time = result
        print(f"\nTotal execution time: {execution_time:.6f} seconds")

        append_execution_time(
            execution_time,
            method="ANOVA - OOP",
            computer_name="Windows Ryzen 9 5900x 32GB"
        )